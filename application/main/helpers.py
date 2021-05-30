from openpyxl import load_workbook
from datetime import datetime as dt
import re
import datetime
import os

from application import current_app as app
from application import db
from flask_login import current_user
from application.main.models import Balance, Credit

def check_db(data_input, signal):

    if signal:
        app.logger.info('INSERTING MOVS ROWS IN DB')
        query = Balance.query.filter(Balance.user_id == current_user.id).all()
        all_db = [[line.bank, line.account_n, line.timestamp, line.detail ,line.flow, line.bal] for line in query]
        data_submitted = []
        for row_data in data_input[::-1]:
            if all_db:
                if row_data in all_db:
                    pass
                else:
                    account = Balance(
                        bank= row_data[0],
                        account_n= row_data[1],
                        timestamp =  row_data[2],
                        detail =  row_data[3],
                        flow =  row_data[4],
                        bal= row_data[5],
                        user_id= current_user.id
                    )
                    db.session.add(account)
                    data_submitted.append([row_data[0],
                        row_data[1],
                        row_data[2],
                        row_data[3],
                        row_data[4],
                        row_data[5]]
                    )
            else:
                account = Balance(
                    bank= row_data[0],
                    account_n= row_data[1],
                    timestamp =  row_data[2],
                    detail =  row_data[3],
                    flow =  row_data[4],
                    bal= row_data[5],
                    user_id= current_user.id
                )
                db.session.add(account)
                data_submitted.append([row_data[0],
                    row_data[1],
                    row_data[2],
                    row_data[3],
                    row_data[4],
                    row_data[5]]
                )
    else:
        app.logger.info('INSERTING CREDIT ROWS IN DB')
        query = Credit.query.filter(Credit.user_id == current_user.id).all()
        all_db = None
        if query:
            # bancos, card, card_number, expiration, date, creditor, share, amount_ars, amount_usd
            all_db = [[line.bank, line.card, line.card_number, line.expiration, line.timestamp.date(), line.creditor, line.share, line.ars , float(line.usd) if line.usd != '0' else 0] for line in query]
            # all_db = [[line.bank, line.timestamp.date(), line.creditor, line.share, line.ars , float(line.usd) if line.usd != '0' else 0] for line in query]

        data_submitted = []
        for row_data in data_input[::-1]:
            if all_db:
                # not insert if row from input data is already in db
                if row_data in all_db:
                    pass
                else:
                    account = Credit(
                        user_id= current_user.id,
                        bank= row_data[0],
                        card = row_data[1],
                        card_number = row_data[2],
                        expiration = row_data[3],
                        timestamp =  row_data[4],
                        creditor= row_data[5],
                        share= row_data[6],
                        ars =  row_data[7],
                        usd =  row_data[8],
                    )
                    db.session.add(account)
                    data_submitted.append(
                        [
                            row_data[0],
                            row_data[1],
                            row_data[2],
                            row_data[3],
                            row_data[4],
                            row_data[5],
                            row_data[6],
                            row_data[7],
                            row_data[8],
                        ]
                    )
            else:
                account = Credit(
                    user_id= current_user.id,
                    bank= row_data[0],
                    card = row_data[1],
                    card_number = row_data[2],
                    expiration = row_data[3],
                    timestamp =  row_data[4],
                    creditor= row_data[5],
                    share= row_data[6],
                    ars =  row_data[7],
                    usd =  row_data[8],
                )
                db.session.add(account)
                data_submitted.append(
                    [
                        row_data[0],
                        row_data[1],
                        row_data[2],
                        row_data[3],
                        row_data[4],
                        row_data[5],
                        row_data[6],
                        row_data[7],
                        row_data[8],
                    ]
                )
    
    db.session.commit()
    return data_submitted


def parse_data(account_num=None, table=None):


    app.logger.info('PROCESSING DATA')
    date = []

    bancos = []
    cuentas = []
    description = []
    balance = []
    flow = []

    creditor = []
    share = []
    amount_ars = []
    amount_usd = []
    card = []
    card_number = []
    expiration = []
    
    if account_num:
        if len(account_num) == 13:
            banco = 'BAPRO'
        elif len(account_num) == 11:
            banco = 'BBVA'
        elif len(account_num) == 12:
            banco = 'Santander'
        bancos += [banco for i in range(len(table['Fecha']))]
        cuentas += [account_num for i in range(len(table['Fecha']))]
            
    # create column with dates if exists
    date += [date for date in table['Fecha'] if date != 'None']

    if 'Card' in table.keys():
        card += table['Card']

    if 'Card_Number' in table.keys():
        card_number += table['Card_Number']

    if 'Expiration' in table.keys():
        expiration += table['Expiration']  

    if 'Concepto' in table.keys():
        description += table['Concepto']
    elif 'Descripción' in table.keys():
        description += table['Descripción']
        # Consolidate in flow the expenses registered in other accounts
        for row in range(len(table['Descripción'])):
            if 'LEY 25413' in table['Descripción'][row]:
                table['Cuenta_sueldo'][row] = table['Importe_cuenta_corriente_pesos'][row]
    elif 'Establecimiento' in table.keys():
        for row in table['Establecimiento']:
            if row != 'None':
                creditor.append(row)
                # add the credit cards bank
                bancos.append('BBVA')
    

    if 'Importe' in table.keys():
        flow += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Importe']]
    elif 'Cuenta_sueldo' in table.keys():
        flow += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Cuenta_sueldo']]
    elif 'Cuota' in table.keys():
        for cuota in table['Cuota']:
            if cuota == '/':
                share.append('')
            elif cuota == 'None':
                pass
            else:
                share.append(cuota)

    if 'Saldo' in table.keys():
        balance += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Saldo']]
    if 'Saldo_pesos' in table.keys():
        balance += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Saldo_pesos']]
    if 'Importe_en_$_' in table.keys():
        for i, j in zip(table['Importe_en_$_'], table['Establecimiento']):
            if j != 'None':
                if i is not None and i != 'None':
                    amount_ars.append(float(i.replace('.','').replace(',','.')))
                else:
                    amount_ars.append(0)

    if 'Importe_en_U$S' in table.keys():
        for i, j in zip(table['Importe_en_U$S'], table['Establecimiento']):
            if j != 'None':
                if i is not None and i != 'None':
                    amount_usd.append(float(i.replace('.','').replace(',','.')))
                else:
                    amount_usd.append(0)

    if balance:
        all_data = bancos, cuentas, date, description, flow, balance
        signal = 1
    else:
        all_data = bancos, card, card_number, expiration, date, creditor, share, amount_ars, amount_usd
        signal = 0

    data_output = []
    skip = []


    for column in range(len(all_data[0])):
        row = []
        for line in all_data:
            # save the place of the list with the element 'TRASPASO'
            if 'TRASPASO' in str(line[column]):
                skip.append(column)
            row.append(line[column])
        data_output.append(row)

    # erase the lists with the element 'TRASPASO' and update the index
    for i, erase in enumerate(skip):
        data_output.remove(data_output[erase-i])


    return check_db(data_output, signal)
    # import pandas as pd
    # print(pd.DataFrame(data_output))


def load_movs(loaded_file):
    app.logger.info('READING XLSX')

    workbook = load_workbook(filename=loaded_file, data_only=True)
    sheet = workbook.active
    app.logger.info(f'Sheet Total Rows: {sheet.max_row}\nSheet Total Columns: {sheet.max_column}')

    # patterns and regex
    p_account = "cuenta(?= única|:)"
    p_account_num = "\d+-+\d+/+\d"
    regex_account = re.compile(p_account, flags = re.IGNORECASE)
    regex_account_num = re.compile(p_account_num, flags = re.IGNORECASE)

    # get account num
    for i in range(1,10):
        for j in range(1,40):
            if re.search(regex_account, str(sheet.cell(row=j, column=i).value)):
                account_num = re.search(regex_account_num, str(sheet.cell(row=j, column=i).value)).group()
                app.logger.info(f'Account num.: {account_num}')

    # search for table vertex
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        row_list = [str(col).lower() if col != None else None for col in row]
        if 'fecha' in row_list:
            for col in row_list:
                if col != None and 'saldo' in col:
                    apex_r = i
                    apex_c = row_list.index('fecha')
                    col_len = sum(elem is not None for elem in row_list)

    # table row length
    row_len = 0
    for row in sheet.iter_rows(values_only=True):
        if isinstance(row[apex_c], dt):
            row_len= row_len+1
        else:
            try:
                dt.strptime(str(row[apex_c]), '%d/%m/%Y')
                row_len= row_len+1
            except:
                pass

    app.logger.info(f'Table Rows: {row_len}\nTable Columns: {col_len}')
    table = dict()

    # populate the table
    header = []
    for i, row in enumerate(list(sheet.rows)[apex_r:(apex_r + row_len + 1)]):
        if i == 0:
            for j, col in enumerate(row[apex_c : col_len + 1]):
                if col.value is not None and ' ' in col.value:
                    col.value = col.value.replace(' ','_')
                header.append(col.value)
                table[col.value] = []
        if i > 0:
            for j, col in enumerate(row[apex_c : col_len + 1]):
                try:
                    table[header[j]].append(dt.strptime(col.value, '%d/%m/%Y').date())
                except:
                    cell = " ".join(str(col.value).strip().split())
                    table[header[j]].append(cell)

    os.remove(loaded_file)

    return parse_data(account_num, table)



def load_credit(loaded_file, card=None, card_number=None, expiration=None):

    workbook = load_workbook(filename=loaded_file, data_only=True)
    sheet = workbook.active
    app.logger.info(f'Sheet Rows: {sheet.max_row}, Sheet Columns: {sheet.max_column}')

    # table vertex -> 1st row (beyond header) and 1st column
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        row_list = [str(col).lower() if col != None else None for col in row]
        if 'fecha' in row_list:
            for col in row_list:
                # avoid header row and columns
                if col != None and 'u$s' in col:
                    apex_row = i
                    apex_col = row_list.index('fecha')
                    col_len = sum(elem is not None for elem in row_list)

    # num of rows above header
    above = 0
    for i in range(len(list(sheet.rows))):
        above += 1
        if list(sheet.rows)[i][1].value == 'Fecha':
            break

    # table row length -> num. of rows with dates
    row_len = 0
    for i in range( above, len(list(sheet.rows)) ):
        row_len += 1
        if list(sheet.rows)[i][1].value == None and list(sheet.rows)[i + 1][1].value == None:
            break

    app.logger.info(f'First Row: {apex_row}, First Column: {apex_col}, Num. of Cols: {col_len}, Num. of Rows: {row_len}')

    table = dict()
    header = []
    # retrieve data of each column
    for i, row in enumerate(list(sheet.rows)[apex_row:(apex_row + row_len)]):
        # get keys from first row
        if i == 0:
            for j, col in enumerate(row[apex_col : col_len + 1]):
                if col.value is not None and ' ' in col.value:
                    col.value = col.value.replace(' ','_')
                header.append(col.value)
                table[col.value] = []
        if i > 1:
            for j, col in enumerate(row[apex_col : col_len + 1]):
                try:
                    table[header[j]].append(dt.strptime(col.value, '%d/%m/%Y').date())
                except:
                    table[header[j]].append(str(col.value).strip())

    table['Card'] = [card for i in range(len(table['Fecha']))]
    table['Card_Number'] = [card_number for i in range(len(table['Fecha']))]

    if expiration:
        date = dt.strptime(expiration, '%Y-%m-%d').strftime('%m/%d/%Y')
        table['Expiration'] = [dt.strptime(date, '%m/%d/%Y').date() for i in range(len(table['Fecha']))]
    else:
        table['Expiration'] = [None for i in range(len(table['Fecha']))]


    return parse_data(table=table)