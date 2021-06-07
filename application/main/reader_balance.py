from openpyxl import load_workbook
from datetime import datetime as dt
import re
import os

from application import current_app as app
from application import db
from flask_login import current_user
from application.main.models import Balance

def balance_db(data_input):

    app.logger.info('INSERTING MOVS ROWS IN DB')
    query = Balance.query.filter(Balance.user_id == current_user.id).all()
    all_db = [[line.bank, line.account_n, line.timestamp, line.detail ,line.flow, line.bal] for line in query]
    data_submitted = []
    for row_data in data_input[::-1]:
        if all_db:
            if row_data in all_db:
                pass
            else:
                account = Balance(
                    bank= row_data[0],
                    account_n= row_data[1],
                    timestamp =  row_data[2],
                    detail =  row_data[3],
                    flow =  row_data[4],
                    bal= row_data[5],
                    user_id= current_user.id
                )
                db.session.add(account)
                data_submitted.append([row_data[0],
                    row_data[1],
                    row_data[2],
                    row_data[3],
                    row_data[4],
                    row_data[5]]
                )
        else:
            account = Balance(
                bank= row_data[0],
                account_n= row_data[1],
                timestamp =  row_data[2],
                detail =  row_data[3],
                flow =  row_data[4],
                bal= row_data[5],
                user_id= current_user.id
            )
            db.session.add(account)
            data_submitted.append([row_data[0],
                row_data[1],
                row_data[2],
                row_data[3],
                row_data[4],
                row_data[5]]
            )
    
    db.session.commit()
    return data_submitted


def parse_data(account_num=None, table=None):


    app.logger.info('PROCESSING DATA')
    date = []

    bancos = []
    cuentas = []
    description = []
    balance = []
    flow = []
    
    # its a cash flow balance
    if account_num:
        if len(account_num) == 13:
            banco = 'BAPRO'
        elif len(account_num) == 11:
            banco = 'BBVA'
        elif len(account_num) == 12:
            banco = 'Santander'
        bancos += [banco for i in range(len(table['Fecha']))]
        cuentas += [account_num for i in range(len(table['Fecha']))]
            
    # create column with dates if exists
    date += [date for date in table['Fecha'] if date != 'None']

    if 'Concepto' in table.keys():
        description += table['Concepto']
    elif 'Descripción' in table.keys():
        description += table['Descripción']
        # Consolidate in flow the expenses registered in other accounts
        for row in range(len(table['Descripción'])):
            if 'LEY 25413' in table['Descripción'][row]:
                table['Cuenta_sueldo'][row] = table['Importe_cuenta_corriente_pesos'][row]
    

    if 'Importe' in table.keys():
        flow += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Importe']]
    elif 'Cuenta_sueldo' in table.keys():
        flow += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Cuenta_sueldo']]

    if 'Saldo' in table.keys():
        balance += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Saldo']]
    if 'Saldo_pesos' in table.keys():
        balance += [float(num.replace('.','').replace(',','.')) if num is not None else 0 for num in table['Saldo_pesos']]

    all_data = bancos, cuentas, date, description, flow, balance

    data_output = []
    skip = []

    for column in range(len(all_data[0])):
        row = []
        for line in all_data:
            # save the place of the list with the element 'TRASPASO'
            if 'TRASPASO' in str(line[column]):
                skip.append(column)
            row.append(line[column])
        data_output.append(row)

    # erase the lists with the element 'TRASPASO' and update the index
    for i, erase in enumerate(skip):
        data_output.remove(data_output[erase-i])

    # import pandas as pd
    # print(pd.DataFrame(data_output))
    return balance_db(data_output)


def load_movs(loaded_file):
    app.logger.info('READING XLSX')

    workbook = load_workbook(filename=loaded_file, data_only=True)
    sheet = workbook.active
    app.logger.info(f'Sheet Total Rows: {sheet.max_row}\nSheet Total Columns: {sheet.max_column}')

    # patterns and regex
    p_account = "cuenta(?= única|:)"
    p_account_num = "\d+-+\d+/+\d"
    regex_account = re.compile(p_account, flags = re.IGNORECASE)
    regex_account_num = re.compile(p_account_num, flags = re.IGNORECASE)

    # get account num
    for i in range(1,10):
        for j in range(1,40):
            if re.search(regex_account, str(sheet.cell(row=j, column=i).value)):
                account_num = re.search(regex_account_num, str(sheet.cell(row=j, column=i).value)).group()
                app.logger.info(f'Account num.: {account_num}')

    # search for table vertex
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        row_list = [str(col).lower() if col != None else None for col in row]
        if 'fecha' in row_list:
            for col in row_list:
                if col != None and 'saldo' in col:
                    apex_r = i
                    apex_c = row_list.index('fecha')
                    col_len = sum(elem is not None for elem in row_list)

    # table row length
    row_len = 0
    for row in sheet.iter_rows(values_only=True):
        if isinstance(row[apex_c], dt):
            row_len= row_len+1
        else:
            try:
                dt.strptime(str(row[apex_c]), '%d/%m/%Y')
                row_len= row_len+1
            except:
                pass

    app.logger.info(f'Table Rows: {row_len}\nTable Columns: {col_len}')
    table = dict()

    # populate the table
    header = []
    for i, row in enumerate(list(sheet.rows)[apex_r:(apex_r + row_len + 1)]):
        if i == 0:
            for j, col in enumerate(row[apex_c : col_len + 1]):
                if col.value is not None and ' ' in col.value:
                    col.value = col.value.replace(' ','_')
                header.append(col.value)
                table[col.value] = []
        if i > 0:
            for j, col in enumerate(row[apex_c : col_len + 1]):
                try:
                    table[header[j]].append(dt.strptime(col.value, '%d/%m/%Y').date())
                except:
                    cell = " ".join(str(col.value).strip().split())
                    table[header[j]].append(cell)

    os.remove(loaded_file)

    return parse_data(account_num=account_num, table=table)



# def load_credit(loaded_file):
    
#     file_name = os.path.split(loaded_file)[-1]
#     pattern = "([12]\d{3}-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01]))"
#     try:
#         close_date = re.search(pattern, file_name).group()
#     except:
#         close_date = None

#     workbook = load_workbook(filename=loaded_file, data_only=True)
#     sheet = workbook.active
#     app.logger.info(f'Sheet Rows: {sheet.max_row}, Sheet Columns: {sheet.max_column}')

#     # table vertex -> 1st row (beyond header) and 1st column
#     for i, row in enumerate(sheet.iter_rows(values_only=True)):
#         row_list = [str(col).lower() if col != None else None for col in row]
#         if 'fecha' in row_list:
#             for col in row_list:
#                 # avoid header row and columns
#                 if col != None and 'u$s' in col:
#                     apex_row = i
#                     apex_col = row_list.index('fecha')
#                     col_len = sum(elem is not None for elem in row_list)

#     # num of rows above header
#     above = 0
#     for i in range(len(list(sheet.rows))):
#         above += 1
#         if list(sheet.rows)[i][1].value == 'Fecha':
#             break

#     # table row length -> num. of rows with dates
#     row_len = 0
#     for i in range( above, len(list(sheet.rows)) ):
#         row_len += 1
#         if list(sheet.rows)[i][1].value == None and list(sheet.rows)[i + 1][1].value == None:
#             break

#     app.logger.info(f'First Row: {apex_row}, First Column: {apex_col}, Num. of Cols: {col_len}, Num. of Rows: {row_len}')

#     table = dict()
#     header = []
#     # retrieve data of each column
#     for i, row in enumerate(list(sheet.rows)[apex_row:(apex_row + row_len)]):
#         # get keys from first row
#         if i == 0:
#             for j, col in enumerate(row[apex_col : col_len + 1]):
#                 if col.value is not None and ' ' in col.value:
#                     col.value = col.value.replace(' ','_')
#                 header.append(col.value)
#                 table[col.value] = []
#         if i > 1:
#             for j, col in enumerate(row[apex_col : col_len + 1]):
#                 try:
#                     table[header[j]].append(dt.strptime(col.value, '%d/%m/%Y').date())
#                 except:
#                     table[header[j]].append(str(col.value).strip())

#     os.remove(loaded_file)

#     return parse_data(table=table, close_date=close_date)